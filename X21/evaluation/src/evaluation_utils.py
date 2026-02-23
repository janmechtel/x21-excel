"""Utility helpers for comparing evaluation workbooks."""

import datetime
import os

import openpyxl


def datetime_to_float(dt):
    """Convert a datetime to Excel serial float."""
    excel_start_date = datetime.datetime(1899, 12, 30)
    delta = dt - excel_start_date
    return delta.days + delta.seconds / 86400.0


def transform_value(v):
    """Normalize cell values for comparison."""
    if isinstance(v, (int, float)):
        v = round(float(v), 2)
    elif isinstance(v, datetime.time):
        v = str(v)[:-3]
    elif isinstance(v, datetime.datetime):
        v = round(datetime_to_float(v), 0)
    elif isinstance(v, str):
        try:
            v = round(float(v), 2)
        except ValueError:
            pass
    return v


def compare_cell_value(v1, v2):
    """Compare two cell values with normalization."""
    v1 = transform_value(v1)
    v2 = transform_value(v2)
    if (v1 == "" and v2 is None) or (v1 is None and v2 == ""):
        return True
    if (v1 == "" and v2 == "") or (v1 is None and v2 is None):
        return True
    if type(v1) is not type(v2):
        # print(type(v1), type(v2))
        return False
    if v1 == v2:
        return True
    else:
        return False


def _get_color_rgb(color) -> str:
    """Extract RGB value from color object, defaulting to '00000000' if not a string."""
    if color and isinstance(color.rgb, str):
        return color.rgb
    return "00000000"


def _compare_colors(color1, color2) -> bool:
    """Compare two colors using only last 6 characters (RGB), ignoring alpha channel."""
    rgb1 = _get_color_rgb(color1)
    rgb2 = _get_color_rgb(color2)
    return rgb1[-6:] == rgb2[-6:]


def compare_fill_color(fill1, fill2) -> bool:
    """Compare fill colors between two cells."""
    return _compare_colors(fill1.fgColor, fill2.fgColor) and _compare_colors(
        fill1.bgColor, fill2.bgColor
    )


def compare_font_color(font_gt, font_proc) -> bool:
    """Compare font colors between two cells."""
    return _compare_colors(font_gt.color, font_proc.color)


def col_num2name(n):
    """Convert a column number to an Excel column name"""
    name = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        name = chr(65 + remainder) + name
    return name


def col_name2num(name):
    """Convert an Excel column name to a column number"""
    num = 0
    for c in name:
        num = num * 26 + (ord(c) - ord("A") + 1)
    return num


def parse_cell_range(range_str):
    """Parse a range string like 'A1:AB12'"""
    start_cell, end_cell = range_str.split(":")
    start_col, start_row = "", ""
    for char in start_cell:
        if char.isdigit():
            start_row += char
        else:
            start_col += char

    end_col, end_row = "", ""
    for char in end_cell:
        if char.isdigit():
            end_row += char
        else:
            end_col += char

    return (col_name2num(start_col), int(start_row)), (
        col_name2num(end_col),
        int(end_row),
    )


def generate_cell_names(range_str):
    """Generate a list of all cell names in the specified range"""
    if ":" not in range_str:
        return [range_str]
    (start_col, start_row), (end_col, end_row) = parse_cell_range(range_str)
    columns = [col_num2name(i) for i in range(start_col, end_col + 1)]
    cell_names = [
        f"{col}{row}" for col in columns for row in range(start_row, end_row + 1)
    ]
    return cell_names


def cell_level_compare(wb_gt, wb_proc, sheet_name, cell_range):
    """Compare cells and return (passed, message, list_of_mismatches)."""
    if sheet_name not in wb_proc:
        return False, "worksheet not found", []
    ws_gt = wb_gt[sheet_name]
    ws_proc = wb_proc[sheet_name]

    cell_names = generate_cell_names(cell_range)
    mismatches = []

    for cell_name in cell_names:
        cell_gt = ws_gt[cell_name]
        cell_proc = ws_proc[cell_name]

        if not compare_cell_value(cell_gt.value, cell_proc.value):
            mismatches.append(
                {
                    "sheet": sheet_name,
                    "cell": cell_gt.coordinate,
                    "expected": cell_gt.value,
                    "actual": cell_proc.value,
                }
            )

    if mismatches:
        msg = f"{len(mismatches)} cell(s) differ"
        return False, msg, mismatches

    return True, "", []


def compare_workbooks(gt_file, proc_file, instruction_type, answer_position):
    """
    Compare workbooks and return (passed, message, all_mismatches).

    all_mismatches is a list of dicts with: sheet, cell, expected, actual
    """
    if not os.path.exists(proc_file):
        return False, "File not exist", []
    # Open workbooks
    try:
        wb_gt = openpyxl.load_workbook(filename=gt_file, data_only=True)
        wb_proc = openpyxl.load_workbook(filename=proc_file, data_only=True)
    except Exception as e:
        return False, str(e), []

    # Initialize report
    all_mismatches = []

    sheet_cell_ranges = answer_position.split(",")
    result_list = []
    for sheet_cell_range in sheet_cell_ranges:
        if "!" in sheet_cell_range:
            sheet_name, cell_range = sheet_cell_range.split("!")
            sheet_name = sheet_name.lstrip("'").rstrip("'")
        else:
            sheet_name = wb_gt.sheetnames[0]
            cell_range = sheet_cell_range

        # process sheet_name and cell_range
        sheet_name = sheet_name.lstrip("'").rstrip("'")
        cell_range = cell_range.lstrip("'").rstrip("'")

        result, msg, mismatches = cell_level_compare(
            wb_gt, wb_proc, sheet_name, cell_range
        )
        result_list.append(result)
        all_mismatches.extend(mismatches)

    passed = all(result_list)
    summary = "" if passed else f"{len(all_mismatches)} cell mismatch(es)"
    return passed, summary, all_mismatches


def evaluate_assertions(workbook_path: str, assertions: list):
    """
    Evaluate assertion list against a workbook.

    Assertions support:
      - equalsText
      - equalsNumber
      - equalsFormat
    """
    if not os.path.exists(workbook_path):
        return False, "File not exist", []

    try:
        wb = openpyxl.load_workbook(filename=workbook_path, data_only=True)
    except Exception as e:
        return False, str(e), []

    mismatches = []
    for assertion in assertions or []:
        sheet_name = assertion.get("sheet") or wb.sheetnames[0]
        cell_ref = assertion.get("cell")
        if not cell_ref:
            continue

        if sheet_name not in wb:
            mismatches.append(
                {
                    "sheet": sheet_name,
                    "cell": cell_ref,
                    "expected": assertion,
                    "actual": None,
                }
            )
            continue

        cell_value = wb[sheet_name][cell_ref].value

        if "equalsText" in assertion:
            expected = assertion.get("equalsText", "")
            actual = "" if cell_value is None else str(cell_value)
            if actual != expected:
                mismatches.append(
                    {
                        "sheet": sheet_name,
                        "cell": cell_ref,
                        "expected": expected,
                        "actual": actual,
                    }
                )
        elif "equalsNumber" in assertion:
            expected = assertion.get("equalsNumber")
            if not compare_cell_value(expected, cell_value):
                mismatches.append(
                    {
                        "sheet": sheet_name,
                        "cell": cell_ref,
                        "expected": expected,
                        "actual": cell_value,
                    }
                )
        elif "equalsFormat" in assertion:
            expected = assertion.get("equalsFormat", "")
            actual = wb[sheet_name][cell_ref].number_format
            if actual != expected:
                mismatches.append(
                    {
                        "sheet": sheet_name,
                        "cell": cell_ref,
                        "expected": expected,
                        "actual": actual,
                    }
                )

    if mismatches:
        return False, f"{len(mismatches)} assertion(s) failed", mismatches

    return True, "", []
